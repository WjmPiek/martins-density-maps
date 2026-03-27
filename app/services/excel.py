from openpyxl import load_workbook

from ..constants import EXPORT_COLUMNS, UPLOAD_COLUMNS
from ..utils.helpers import build_full_address, normalize_float, normalize_text
from .geocoding import geocode_address

LEGACY_UPLOAD_COLUMNS = [
    'MF File',
    'Deceased Name',
    'Deceased Surname',
    'DOD',
    'Deceased Address',
    'Address',
    'City',
    'Province',
    'Postal Code',
    'Country',
    'Church Name',
    'Church Street Address',
    'Church City',
    'Church Province',
    'Church Postal Code',
    'Church Country',
    'Church Address',
    'Pastor Name',
    'Church Mobile Number',
    'Next of Kin Name',
    'Next of Kin Surname',
    'Relationship',
    'Contact Number',
]

LEGACY_EXPORT_COLUMNS = [
    'MF File',
    'Deceased Name',
    'Deceased Surname',
    'DOD',
    'Deceased Address',
    'Address',
    'City',
    'Province',
    'Postal Code',
    'Country',
    'Full Address',
    'Latitude',
    'Longitude',
    'Weight',
    'Church Name',
    'Church Street Address',
    'Church City',
    'Church Province',
    'Church Postal Code',
    'Church Country',
    'Church Address',
    'Pastor Name',
    'Church Mobile Number',
    'Next of Kin Name',
    'Next of Kin Surname',
    'Relationship',
    'Contact Number',
]

SUPPORTED_HEADERS = [
    UPLOAD_COLUMNS,
    EXPORT_COLUMNS,
    LEGACY_UPLOAD_COLUMNS,
    LEGACY_EXPORT_COLUMNS,
]


def _match_header(header_row):
    for supported in SUPPORTED_HEADERS:
        if header_row[: len(supported)] == supported:
            return supported
    return None



def parse_upload(file_storage):
    file_storage.stream.seek(0)
    wb = load_workbook(file_storage.stream, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = [normalize_text(cell) for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    matched_header = _match_header(header_row)
    if not matched_header:
        raise ValueError('Workbook columns do not match the required Martins template.')

    records = []
    warnings = []
    seen_mf = set()

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = dict(zip(matched_header, row[: len(matched_header)]))
        mf_file = normalize_text(data.get('MF File'))

        if not any(normalize_text(v) for v in data.values()):
            continue
        if not mf_file:
            warnings.append(f'Row {idx}: missing MF File and skipped.')
            continue
        if mf_file in seen_mf:
            warnings.append(f"Row {idx}: duplicate MF File '{mf_file}' skipped.")
            continue
        seen_mf.add(mf_file)

        address = normalize_text(data.get('Address'))
        city = normalize_text(data.get('City'))
        province = normalize_text(data.get('Province'))
        postal_code = normalize_text(data.get('Postal Code'))
        country = normalize_text(data.get('Country')) or 'South Africa'

        deceased_address = normalize_text(data.get('Deceased Address')) or build_full_address(
            address,
            city,
            province,
            postal_code,
            country,
        )
        full_address = normalize_text(data.get('Full Address')) or build_full_address(
            address,
            city,
            province,
            postal_code,
            country,
        )

        latitude = normalize_float(data.get('Latitude'))
        longitude = normalize_float(data.get('Longitude'))
        if latitude is None or longitude is None:
            latitude, longitude = geocode_address(full_address)

        church_street_address = normalize_text(data.get('Church Street Address'))
        church_city = normalize_text(data.get('Church City'))
        church_province = normalize_text(data.get('Church Province'))
        church_postal_code = normalize_text(data.get('Church Postal Code'))
        church_country = normalize_text(data.get('Church Country')) or 'South Africa'
        church_address = normalize_text(data.get('Church Address')) or build_full_address(
            church_street_address,
            church_city,
            church_province,
            church_postal_code,
            church_country,
        )

        records.append(
            {
                'mf_file': mf_file,
                'deceased_name': normalize_text(data.get('Deceased Name')),
                'deceased_surname': normalize_text(data.get('Deceased Surname')),
                'dod': normalize_text(data.get('DOD')),
                'deceased_address': deceased_address,
                'address': address,
                'city': city,
                'province': province,
                'postal_code': postal_code,
                'country': country,
                'full_address': full_address,
                'latitude': latitude,
                'longitude': longitude,
                'weight': normalize_float(data.get('Weight')) or 1.0,
                'church_name': normalize_text(data.get('Church Name')),
                'church_street_address': church_street_address,
                'church_city': church_city,
                'church_province': church_province,
                'church_postal_code': church_postal_code,
                'church_country': church_country,
                'church_address': church_address,
                'pastor_name': normalize_text(data.get('Pastor Name')),
                'church_mobile_number': normalize_text(data.get('Church Mobile Number')),
                'next_of_kin_name': normalize_text(data.get('Next of Kin Name')),
                'next_of_kin_surname': normalize_text(data.get('Next of Kin Surname')),
                'relationship': normalize_text(data.get('Relationship')),
                'contact_number': normalize_text(data.get('Contact Number')),
            }
        )

    return records, warnings
