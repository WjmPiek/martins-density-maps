#!/usr/bin/env python3
import http.server, ssl, os, sys, json
from urllib.parse import urlparse
from openpyxl import load_workbook


PORT = 8443
CERT = "localhost.pem"
KEY  = "localhost-key.pem"

class Handler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        self.send_header("Cache-Control", "no-store")
        super().end_headers()

    def _send_json(self, payload, status=200):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_GET(self):
        parsed = urlparse(self.path)
        if parsed.path == "/data.json":
            xlsx_path = os.path.join(os.getcwd(), "martins_density_map_data.xlsx")
            if not os.path.exists(xlsx_path):
                return self._send_json({"error": "Workbook not found", "file": "martins_density_map_data.xlsx"}, status=404)
            try:
                wb = load_workbook(xlsx_path, data_only=True)
                ws = wb[wb.sheetnames[0]]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return self._send_json({"rows": []})
                headers = [str(v).strip() if v is not None else "" for v in rows[0]]
                data = []
                for row in rows[1:]:
                    item = {}
                    for i, header in enumerate(headers):
                        if not header:
                            continue
                        value = row[i] if i < len(row) else ""
                        item[header] = "" if value is None else value
                    data.append(item)
                return self._send_json({"rows": data})
            except Exception as e:
                return self._send_json({"error": str(e)}, status=500)
        return super().do_GET()

def main():
    if not (os.path.exists(CERT) and os.path.exists(KEY)):
        print("Missing certificate files:")
        print(f"  {CERT}")
        print(f"  {KEY}")
        print()
        print("Run the PowerShell script run_https_windows.ps1 to generate them.")
        print("Or generate with OpenSSL:")
        print('  openssl req -x509 -newkey rsa:2048 -nodes -keyout localhost-key.pem -out localhost.pem -days 365 -subj "/CN=localhost"')
        sys.exit(1)

    httpd = http.server.ThreadingHTTPServer(("localhost", PORT), Handler)
    ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_SERVER)
    ctx.load_cert_chain(certfile=CERT, keyfile=KEY)
    httpd.socket = ctx.wrap_socket(httpd.socket, server_side=True)

    print(f"Serving HTTPS on https://localhost:{PORT}/heatmap.html")
    print("Press Ctrl+C to stop.")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass

if __name__ == "__main__":
    main()
