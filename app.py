import os
from datetime import datetime
from functools import wraps
from io import BytesIO

import pandas as pd
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_sqlalchemy import SQLAlchemy
from geopy.geocoders import Nominatim
from openpyxl import Workbook, load_workbook
from sqlalchemy import func
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


app = Flask(__name__)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
LOGO_PATH = os.path.join(BASE_DIR, "static", "img", "martins-logo.png")
TEMPLATE_EXPORT = os.path.join(BASE_DIR, "martins_density_map_data.xlsx")
TEMPLATE_DOWNLOAD_PATH = os.path.join(
    BASE_DIR, "static", "templates", "martins_density_map_template.xlsx"
)

UPLOAD_TEMPLATE_COLUMNS = [
    "MF File",
    "Deceased Name",
    "Deceased Surname",
    "DOD",
    "Address",
    "City",
    "Province",
    "Country",
    "Next of Kin Name",
    "Next of Kin Surname",
    "Relationship",
    "Contact Number",
]

DERIVED_COLUMNS = ["Full Address", "Latitude", "Longitude", "Weight"]

EXPORT_COLUMNS = [
    "MF File",
    "Deceased Name",
    "Deceased Surname",
    "DOD",
    "Address",
    "City",
    "Province",
    "Country",
    "Full Address",
    "Latitude",
    "Longitude",
    "Weight",
    "Next of Kin Name",
    "Next of Kin Surname",
    "Relationship",
    "Contact Number",
]

PROVINCES = [
    "Eastern Cape",
    "Free State",
    "Gauteng",
    "KwaZulu-Natal",
    "Limpopo",
    "Mpumalanga",
    "North West",
    "Northern Cape",
    "Western Cape",
]

SAMPLE_TEMPLATE_ROWS = [
    {
        "MF File": "MF001",
        "Deceased Name": "John",
        "Deceased Surname": "Smith",
        "DOD": "2024-01-12",
        "Address": "12 Oak Street",
        "City": "Pretoria",
        "Province": "Gauteng",
        "Country": "South Africa",
        "Next of Kin Name": "Mary",
        "Next of Kin Surname": "Smith",
        "Relationship": "Wife",
        "Contact Number": "825551234",
    },
    {
        "MF File": "MF002",
        "Deceased Name": "Peter",
        "Deceased Surname": "Mokoena",
        "DOD": "2024-02-03",
        "Address": "45 Church Rd",
        "City": "Johannesburg",
        "Province": "Gauteng",
        "Country": "South Africa",
        "Next of Kin Name": "Sarah",
        "Next of Kin Surname": "Mokoena",
        "Relationship": "Daughter",
        "Contact Number": "725557788",
    },
    {
        "MF File": "MF003",
        "Deceased Name": "Nomsa",
        "Deceased Surname": "Dlamini",
        "DOD": "2024-03-20",
        "Address": "8 Lagoon Ave",
        "City": "Durban",
        "Province": "KwaZulu-Natal",
        "Country": "South Africa",
        "Next of Kin Name": "Sipho",
        "Next of Kin Surname": "Dlamini",
        "Relationship": "Son",
        "Contact Number": "834448899",
    },
]

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(INSTANCE_DIR, exist_ok=True)

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "change-me-in-render")
app.config["SQLALCHEMY_DATABASE_URI"] = os.environ.get(
    "DATABASE_URL", f"sqlite:///{os.path.join(INSTANCE_DIR, 'app.db')}"
)
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

if app.config["SQLALCHEMY_DATABASE_URI"].startswith("postgres://"):
    app.config["SQLALCHEMY_DATABASE_URI"] = app.config["SQLALCHEMY_DATABASE_URI"].replace(
        "postgres://", "postgresql://", 1
    )

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"
login_manager.login_message_category = "warning"


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(255), unique=True, nullable=False, index=True)
    password_hash = db.Column(db.String(255), nullable=False)
    role = db.Column(db.String(20), nullable=False, default="user")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)

    uploads = db.relationship("Upload", backref="user", lazy=True, cascade="all, delete-orphan")
    records = db.relationship("Record", backref="user", lazy=True, cascade="all, delete-orphan")

    def set_password(self, password: str) -> None:
        self.password_hash = generate_password_hash(password)

    def check_password(self, password: str) -> bool:
        return check_password_hash(self.password_hash, password)

    @property
    def is_admin(self) -> bool:
        return self.role == "admin"


class Upload(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    imported_rows = db.Column(db.Integer, nullable=False, default=0)
    status = db.Column(db.String(30), nullable=False, default="completed")
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)


class Record(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    mf_file = db.Column(db.String(120), nullable=False, index=True)
    deceased_name = db.Column(db.String(120))
    deceased_surname = db.Column(db.String(120))
    dod = db.Column(db.String(50))
    address = db.Column(db.String(255))
    city = db.Column(db.String(120), index=True)
    province = db.Column(db.String(120), index=True)
    country = db.Column(db.String(120))
    full_address = db.Column(db.String(512))
    latitude = db.Column(db.Float)
    longitude = db.Column(db.Float)
    weight = db.Column(db.Float, default=1.0)
    next_of_kin_name = db.Column(db.String(120))
    next_of_kin_surname = db.Column(db.String(120))
    relationship = db.Column(db.String(120))
    contact_number = db.Column(db.String(120))
    created_at = db.Column(db.DateTime, default=datetime.utcnow, nullable=False)
    updated_at = db.Column(db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow, nullable=False)

    __table_args__ = (
        db.UniqueConstraint("user_id", "mf_file", name="uq_user_mf_file"),
    )

    def to_dict(self):
        return {
            "id": self.id,
            "mfFile": self.mf_file,
            "deceasedName": self.deceased_name or "",
            "deceasedSurname": self.deceased_surname or "",
            "dod": self.dod or "",
            "address": self.address or "",
            "city": self.city or "",
            "province": self.province or "",
            "country": self.country or "",
            "fullAddress": self.full_address or "",
            "latitude": self.latitude,
            "longitude": self.longitude,
            "weight": self.weight if self.weight is not None else 1,
            "nextOfKinName": self.next_of_kin_name or "",
            "nextOfKinSurname": self.next_of_kin_surname or "",
            "relationship": self.relationship or "",
            "contactNumber": self.contact_number or "",
            "owner": self.user.name,
            "ownerEmail": self.user.email,
            "updatedAt": self.updated_at.isoformat(),
        }


@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))


def admin_required(view_func):
    @wraps(view_func)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_admin:
            flash("Admin access required.", "danger")
            return redirect(url_for("dashboard"))
        return view_func(*args, **kwargs)

    return wrapped


def bootstrap_admin() -> None:
    admin_email = os.environ.get("ADMIN_EMAIL")
    admin_password = os.environ.get("ADMIN_PASSWORD")
    admin_name = os.environ.get("ADMIN_NAME", "Martins Admin")

    if not admin_email or not admin_password:
        return

    existing = User.query.filter(func.lower(User.email) == admin_email.lower()).first()
    if existing:
        if existing.role != "admin":
            existing.role = "admin"
            db.session.commit()
        return

    user = User(name=admin_name, email=admin_email.lower(), role="admin")
    user.set_password(admin_password)
    db.session.add(user)
    db.session.commit()


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def normalize_float(value):
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None

    text = normalize_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def derive_full_address(address, city, province, country, supplied=None):
    supplied_text = normalize_text(supplied)
    if supplied_text:
        return supplied_text
    return ", ".join(part for part in [address, city, province, country] if part)


def geocode_address(full_address):
    full_address = normalize_text(full_address)
    if not full_address:
        return None, None

    try:
        geolocator = Nominatim(user_agent="martins_density_map")
        location = geolocator.geocode(full_address, timeout=10)
        if location:
            return location.latitude, location.longitude
    except Exception:
        pass

    return None, None


def finalize_record_values(data):
    address = normalize_text(data.get("Address"))
    city = normalize_text(data.get("City"))
    province = normalize_text(data.get("Province"))
    country = normalize_text(data.get("Country"))
    full_address = derive_full_address(
        address,
        city,
        province,
        country,
        supplied=data.get("Full Address"),
    )

    latitude = normalize_float(data.get("Latitude"))
    longitude = normalize_float(data.get("Longitude"))
    weight = normalize_float(data.get("Weight"))

    if (latitude is None or longitude is None) and full_address:
        latitude, longitude = geocode_address(full_address)

    return {
        "mf_file": normalize_text(data.get("MF File")),
        "deceased_name": normalize_text(data.get("Deceased Name")),
        "deceased_surname": normalize_text(data.get("Deceased Surname")),
        "dod": normalize_text(data.get("DOD")),
        "address": address,
        "city": city,
        "province": province,
        "country": country,
        "full_address": full_address,
        "latitude": latitude,
        "longitude": longitude,
        "weight": weight if weight is not None else 1.0,
        "next_of_kin_name": normalize_text(data.get("Next of Kin Name")),
        "next_of_kin_surname": normalize_text(data.get("Next of Kin Surname")),
        "relationship": normalize_text(data.get("Relationship")),
        "contact_number": normalize_text(data.get("Contact Number")),
    }


def parse_upload(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = [
        normalize_text(cell)
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    header_map = {header: idx for idx, header in enumerate(header_row) if header}

    missing_cols = [col for col in UPLOAD_TEMPLATE_COLUMNS if col not in header_map]
    if missing_cols:
        raise ValueError(
            f"Workbook is missing required columns: {', '.join(missing_cols)}"
        )

    records = []
    warnings = []
    seen_mf = set()

    available_columns = [col for col in EXPORT_COLUMNS if col in header_map]

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        data = {
            col: row[header_map[col]] if header_map[col] < len(row) else None
            for col in available_columns
        }

        if not any(normalize_text(v) for v in data.values()):
            continue

        mf_file = normalize_text(data.get("MF File"))
        if not mf_file:
            warnings.append(f"Row {idx}: missing MF File and skipped.")
            continue

        if mf_file in seen_mf:
            warnings.append(f"Row {idx}: duplicate MF File '{mf_file}' in upload.")
        seen_mf.add(mf_file)

        record_values = finalize_record_values(data)
        if not record_values["mf_file"]:
            warnings.append(f"Row {idx}: missing MF File and skipped.")
            continue

        records.append(record_values)

    return records, warnings


def upsert_record(user_id, payload):
    mf_file = normalize_text(payload.get("mfFile") or payload.get("MF File"))
    if not mf_file:
        raise ValueError("MF File is required.")

    record = Record.query.filter_by(user_id=user_id, mf_file=mf_file).first()
    if record is None:
        record = Record(user_id=user_id, mf_file=mf_file)
        db.session.add(record)

    normalized = finalize_record_values(
        {
            "MF File": mf_file,
            "Deceased Name": payload.get("deceasedName") or payload.get("Deceased Name"),
            "Deceased Surname": payload.get("deceasedSurname") or payload.get("Deceased Surname"),
            "DOD": payload.get("dod") or payload.get("DOD"),
            "Address": payload.get("address") or payload.get("Address"),
            "City": payload.get("city") or payload.get("City"),
            "Province": payload.get("province") or payload.get("Province"),
            "Country": payload.get("country") or payload.get("Country"),
            "Full Address": payload.get("fullAddress") or payload.get("Full Address"),
            "Latitude": payload.get("latitude") or payload.get("Latitude"),
            "Longitude": payload.get("longitude") or payload.get("Longitude"),
            "Weight": payload.get("weight") or payload.get("Weight"),
            "Next of Kin Name": payload.get("NextOfKinName") or payload.get("nextOfKinName") or payload.get("Next of Kin Name"),
            "Next of Kin Surname": payload.get("NextOfKinSurname") or payload.get("nextOfKinSurname") or payload.get("Next of Kin Surname"),
            "Relationship": payload.get("relationship") or payload.get("Relationship"),
            "Contact Number": payload.get("contactNumber") or payload.get("Contact Number"),
        }
    )

    record.deceased_name = normalized["deceased_name"]
    record.deceased_surname = normalized["deceased_surname"]
    record.dod = normalized["dod"]
    record.address = normalized["address"]
    record.city = normalized["city"]
    record.province = normalized["province"]
    record.country = normalized["country"]
    record.full_address = normalized["full_address"]
    record.latitude = normalized["latitude"]
    record.longitude = normalized["longitude"]
    record.weight = normalized["weight"]
    record.next_of_kin_name = normalized["next_of_kin_name"]
    record.next_of_kin_surname = normalized["next_of_kin_surname"]
    record.relationship = normalized["relationship"]
    record.contact_number = normalized["contact_number"]
    return record


def build_workbook(records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(EXPORT_COLUMNS)

    for record in records:
        ws.append(
            [
                record.mf_file,
                record.deceased_name,
                record.deceased_surname,
                record.dod,
                record.address,
                record.city,
                record.province,
                record.country,
                record.full_address,
                record.latitude,
                record.longitude,
                record.weight,
                record.next_of_kin_name,
                record.next_of_kin_surname,
                record.relationship,
                record.contact_number,
            ]
        )

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 28)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def build_template_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(UPLOAD_TEMPLATE_COLUMNS)

    for row in SAMPLE_TEMPLATE_ROWS:
        ws.append([row.get(col, "") for col in UPLOAD_TEMPLATE_COLUMNS])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 28)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def dataset_for_user(user):
    query = Record.query
    if not user.is_admin:
        query = query.filter_by(user_id=user.id)

    records = query.order_by(Record.city.asc(), Record.mf_file.asc()).all()
    mapped = sum(1 for r in records if r.latitude is not None and r.longitude is not None)
    provinces = sorted({r.province for r in records if r.province})
    owners = sorted({r.user.name for r in records})

    return {
        "records": [r.to_dict() for r in records],
        "summary": {
            "total": len(records),
            "mapped": mapped,
            "unmapped": len(records) - mapped,
            "provinces": provinces,
            "owners": owners,
        },
    }


@app.route("/upload-page")
@login_required
def upload_page():
    return render_template("upload.html")


@app.route("/download-template")
@login_required
def download_template():
    stream = build_template_workbook()
    return send_file(
        stream,
        as_attachment=True,
        download_name="martins_density_map_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/")
def index():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


@app.route("/register", methods=["GET", "POST"])
def register():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        name = normalize_text(request.form.get("name"))
        email = normalize_text(request.form.get("email")).lower()
        password = request.form.get("password", "")
        confirm = request.form.get("confirm_password", "")

        if not name or not email or not password:
            flash("Name, email, and password are required.", "danger")
        elif password != confirm:
            flash("Passwords do not match.", "danger")
        elif User.query.filter(func.lower(User.email) == email).first():
            flash("An account with that email already exists.", "warning")
        else:
            user = User(name=name, email=email, role="user")
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            login_user(user)
            flash("Welcome to Martins Density Map.", "success")
            return redirect(url_for("dashboard"))

    return render_template("register.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = normalize_text(request.form.get("email")).lower()
        password = request.form.get("password", "")
        user = User.query.filter(func.lower(User.email) == email).first()

        if user and user.check_password(password):
            login_user(user, remember=True)
            flash("Signed in successfully.", "success")
            next_url = request.args.get("next")
            return redirect(next_url or url_for("dashboard"))

        flash("Invalid email or password.", "danger")

    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


@app.route("/dashboard")
@login_required
def dashboard():
    return render_template(
        "dashboard.html",
        provinces=PROVINCES,
        google_maps_api_key=os.environ.get("GOOGLE_MAPS_API_KEY", ""),
    )


@app.route("/admin")
@login_required
@admin_required
def admin():
    user_count = User.query.count()
    record_count = Record.query.count()
    upload_count = Upload.query.count()
    latest_uploads = Upload.query.order_by(Upload.created_at.desc()).limit(10).all()

    return render_template(
        "admin.html",
        user_count=user_count,
        record_count=record_count,
        upload_count=upload_count,
        latest_uploads=latest_uploads,
    )


@app.route("/api/records")
@login_required
def api_records():
    return jsonify(dataset_for_user(current_user))


@app.route("/api/records", methods=["POST"])
@login_required
def api_save_record():
    payload = request.get_json(force=True)

    try:
        record = upsert_record(current_user.id, payload)
        db.session.commit()
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Could not save record."}), 500

    return jsonify({"message": "Record saved.", "record": record.to_dict()})


@app.route("/api/records/<int:record_id>", methods=["DELETE"])
@login_required
def api_delete_record(record_id):
    record = Record.query.get_or_404(record_id)

    if not current_user.is_admin and record.user_id != current_user.id:
        return jsonify({"error": "Not allowed."}), 403

    db.session.delete(record)
    db.session.commit()
    return jsonify({"message": "Record deleted."})


@app.route("/api/upload", methods=["POST"])
@login_required
def api_upload():
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "Select an Excel file to upload."}), 400

    filename = secure_filename(file.filename)

    try:
        imported_rows, warnings = parse_upload(file)

        Record.query.filter_by(user_id=current_user.id).delete()

        for row in imported_rows:
            db.session.add(Record(user_id=current_user.id, **row))

        stored_name = f"{current_user.id}_{int(datetime.utcnow().timestamp())}_{filename}"
        file.stream.seek(0)
        file.save(os.path.join(UPLOAD_DIR, stored_name))

        db.session.add(
            Upload(
                user_id=current_user.id,
                filename=stored_name,
                original_filename=filename,
                imported_rows=len(imported_rows),
                status="completed",
            )
        )

        db.session.commit()

        return jsonify(
            {
                "message": f"Imported {len(imported_rows)} records.",
                "warnings": warnings,
            }
        )
    except ValueError as exc:
        db.session.rollback()
        return jsonify({"error": str(exc)}), 400
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Upload failed."}), 500


@app.route("/upload", methods=["POST"])
@login_required
def upload_excel():
    file = request.files.get("file")
    if not file or not file.filename:
        flash("Please choose a file.", "warning")
        return redirect(url_for("dashboard"))

    filename = file.filename.lower()

    try:
        if filename.endswith(".csv"):
            df = pd.read_csv(file)
        else:
            df = pd.read_excel(file)
    except Exception:
        flash("Could not read the uploaded file. Please use the Martins template.", "danger")
        return redirect(url_for("dashboard"))

    df.columns = [normalize_text(col) for col in df.columns]

    missing_cols = [col for col in UPLOAD_TEMPLATE_COLUMNS if col not in df.columns]
    if missing_cols:
        flash(f"Missing required columns: {', '.join(missing_cols)}", "danger")
        return redirect(url_for("dashboard"))

    try:
        Record.query.filter_by(user_id=current_user.id).delete()

        for _, row in df.iterrows():
            raw = {col: row.get(col) for col in df.columns}

            if not any(normalize_text(value) for value in raw.values()):
                continue

            record_values = finalize_record_values(raw)
            if not record_values["mf_file"]:
                continue

            db.session.add(Record(user_id=current_user.id, **record_values))

        db.session.commit()
        flash("File uploaded successfully.", "success")
    except Exception:
        db.session.rollback()
        flash("Upload failed. Please verify the template columns and data types.", "danger")

    return redirect(url_for("dashboard"))


@app.route("/download/my-data.xlsx")
@login_required
def download_my_data():
    records = (
        Record.query.filter_by(user_id=current_user.id)
        .order_by(Record.city.asc(), Record.mf_file.asc())
        .all()
    )
    stream = build_workbook(records)

    return send_file(
        stream,
        as_attachment=True,
        download_name="my_martins_density_map_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/download/central.xlsx")
@login_required
@admin_required
def download_central():
    records = (
        Record.query.join(User)
        .order_by(User.name.asc(), Record.city.asc(), Record.mf_file.asc())
        .all()
    )
    stream = build_workbook(records)

    return send_file(
        stream,
        as_attachment=True,
        download_name="martins_density_map_data.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.context_processor
def inject_globals():
    return {
        "google_maps_api_key": os.environ.get("GOOGLE_MAPS_API_KEY", ""),
        "logo_path": url_for("static", filename="img/martins-logo.png"),
    }


with app.app_context():
    db.create_all()
    bootstrap_admin()


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)))
